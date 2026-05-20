import io
import logging
from datetime import UTC, datetime
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.sql import ColumnElement, Select

from src.app.services.case.models import Case, CaseStatus, case_experts
from src.app.services.client.models import Client
from src.app.services.user.models import User, UserRole

logger = logging.getLogger(__name__)

BATCH_SIZE = 1000


class CaseExcelExportService:
    def __init__(self, db_session: AsyncSession) -> None:
        self.db = db_session

    async def export_cases_to_excel(
        self,
        user_id: UUID,
        user_role: UserRole,
        company_id: UUID,
        filters: dict[str, Any] | None = None,
    ) -> bytes:
        """Export all cases matching filters to Excel with optimized batched queries."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Все дела"
        self._setup_headers(ws)

        base_query = self._build_base_query(company_id, user_id, user_role, filters)
        total_count = await self._get_total_count(base_query)
        logger.info(f"Exporting {total_count} cases to Excel")

        await self._write_data_batched(ws, base_query, total_count)
        self._apply_formatting(ws)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _build_base_query(
        self,
        company_id: UUID,
        user_id: UUID,
        user_role: UserRole,
        filters: dict[str, Any] | None,
    ) -> Select[Any]:
        """Build optimized query with minimal joins and eager loading."""
        stmt = (
            select(
                Case.id,
                Case.number,
                Case.case_number,
                Case.authority,
                Case.case_type,
                Case.object_type,
                Case.object_address,
                Case.status,
                Case.start_date,
                Case.deadline,
                Case.completion_date,
                Case.cost,
                Case.bank_transfer_amount,
                Case.cash_amount,
                Case.remaining_debt,
                Case.debit,
                Case.plaintiff,
                Case.defendant,
                Case.judge_name,
                Case.remarks,
                Case.created_at,
                Case.updated_at,
                Case.deleted_at,
                Client.name.label("client_name"),
                Client.inn.label("client_inn"),
                Client.email.label("client_email"),
                Client.phone.label("client_phone"),
            )
            .outerjoin(Client, Case.client_id == Client.id)
            .where(Case.deleted_at.is_(None), Case.company_id == company_id)
        )

        if user_role == UserRole.EXPERT:
            stmt = stmt.where(Case.id.in_(select(case_experts.c.case_id).where(case_experts.c.user_id == user_id)))

        if filters:
            stmt = self._apply_filters(stmt, filters)

        return stmt.order_by(Case.created_at.desc())

    def _apply_filters(self, stmt: Select[Any], filters: dict[str, Any]) -> Select[Any]:
        """Apply filters to query."""
        filter_mappings: dict[str, Any] = {
            "status": lambda v: Case.status.in_(v) if isinstance(v, list) else Case.status == v,
            "client_id": lambda v: Case.client_id == v,
            "start_date": lambda v: Case.start_date >= v,
            "end_date": lambda v: Case.start_date <= v,
            "case_type": lambda v: Case.case_type.ilike(f"%{v}%"),
            "object_type": lambda v: Case.object_type.ilike(f"%{v}%"),
            "authority": lambda v: Case.authority.ilike(f"%{v}%"),
            "object_address": lambda v: Case.object_address.ilike(f"%{v}%"),
            "number": lambda v: Case.number == v,
            "case_number": lambda v: Case.case_number == v,
            "min_cost": lambda v: Case.cost >= v,
            "max_cost": lambda v: Case.cost <= v,
            "min_remaining_debt": lambda v: Case.remaining_debt >= v,
            "max_remaining_debt": lambda v: Case.remaining_debt <= v,
        }

        for key, value in filters.items():
            if value is not None and key in filter_mappings:
                condition = filter_mappings[key](value)
                stmt = stmt.where(condition)

        if filters.get("search"):
            search_condition = self._build_search_condition(filters["search"])
            stmt = stmt.outerjoin(Client, Case.client_id == Client.id).where(search_condition)
        return stmt

    def _build_search_condition(self, search_term: str) -> ColumnElement[bool]:
        """Build search condition for multiple fields."""
        pattern = f"%{search_term}%"
        return (
            Case.number.ilike(pattern)
            | Case.case_number.ilike(pattern)
            | Case.authority.ilike(pattern)
            | Case.object_address.ilike(pattern)
            | Case.plaintiff.ilike(pattern)
            | Case.defendant.ilike(pattern)
            | Case.remarks.ilike(pattern)
        )

    async def _get_total_count(self, stmt: Select[Any]) -> int:
        """Get total count for logging purposes."""
        count_stmt = select(func.count()).select_from(stmt.subquery())
        result = await self.db.execute(count_stmt)
        return result.scalar() or 0

    async def _write_data_batched(self, ws: Worksheet, base_stmt: Select[Any], total_count: int) -> None:
        """Write data to worksheet in optimized batches."""
        row_num = 2
        for offset in range(0, max(total_count, 1), BATCH_SIZE):
            stmt = base_stmt.offset(offset).limit(BATCH_SIZE)
            result = await self.db.execute(stmt)
            rows = result.all()

            if not rows:
                break

            for case_row in rows:
                experts = await self._get_experts_for_case(case_row.id)
                experts_str = ", ".join(experts) if experts else ""

                ws.cell(row=row_num, column=1, value=str(case_row.id))
                ws.cell(row=row_num, column=2, value=case_row.number)
                ws.cell(row=row_num, column=3, value=case_row.case_number)
                ws.cell(row=row_num, column=4, value=case_row.client_name or "")
                ws.cell(row=row_num, column=5, value=case_row.client_inn or "")
                ws.cell(row=row_num, column=6, value=case_row.client_email or "")
                ws.cell(row=row_num, column=7, value=case_row.client_phone or "")
                ws.cell(row=row_num, column=8, value=case_row.authority)
                ws.cell(row=row_num, column=9, value=case_row.case_type)
                ws.cell(row=row_num, column=10, value=case_row.object_type)
                ws.cell(row=row_num, column=11, value=case_row.object_address)
                ws.cell(row=row_num, column=12, value=self._format_status(case_row.status))
                ws.cell(row=row_num, column=13, value=self._format_datetime(case_row.start_date))
                ws.cell(row=row_num, column=14, value=self._format_datetime(case_row.deadline))
                ws.cell(row=row_num, column=15, value=self._format_datetime(case_row.completion_date))
                ws.cell(row=row_num, column=16, value=float(case_row.cost) if case_row.cost else 0)
                ws.cell(row=row_num, column=17, value=float(case_row.bank_transfer_amount) if case_row.bank_transfer_amount else 0)
                ws.cell(row=row_num, column=18, value=float(case_row.cash_amount) if case_row.cash_amount else 0)
                ws.cell(row=row_num, column=19, value=float(case_row.remaining_debt) if case_row.remaining_debt else 0)

                ws.cell(row=row_num, column=20, value=float(case_row.debit) if case_row.debit else 0)

                ws.cell(row=row_num, column=21, value=case_row.plaintiff or "")
                ws.cell(row=row_num, column=22, value=case_row.defendant or "")
                ws.cell(row=row_num, column=23, value=case_row.judge_name or "")
                ws.cell(row=row_num, column=24, value=case_row.remarks or "")
                ws.cell(row=row_num, column=25, value=experts_str)
                ws.cell(row=row_num, column=26, value=self._format_datetime(case_row.created_at))
                ws.cell(row=row_num, column=27, value=self._format_datetime(case_row.updated_at))

                row_num += 1

    async def _get_experts_for_case(self, case_id: UUID) -> list[str]:
        """Get expert names for a case in optimized way."""
        stmt = (
            select(User.full_name, User.email)
            .select_from(case_experts)
            .join(User, case_experts.c.user_id == User.id)
            .where(case_experts.c.case_id == case_id)
        )
        result = await self.db.execute(stmt)
        experts = result.all()
        return [f"{e.full_name} ({e.email})" if e.full_name else e.email for e in experts]

    def _format_status(self, status: CaseStatus) -> str:
        """Format case status for display."""
        status_map = {
            CaseStatus.archive: "Архив",
            CaseStatus.in_work: "В работе",
            CaseStatus.debt: "Долг",
            CaseStatus.executed: "Исполнено",
            CaseStatus.withdrawn: "Отозвано",
            CaseStatus.cancelled: "Отменено",
            CaseStatus.fssp: "ФССП",
        }
        return status_map.get(status, status.value)

    def _format_datetime(self, dt: datetime | None) -> str:
        """Format datetime for Excel."""
        if dt is None:
            return ""
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=UTC)
        return dt.strftime("%Y-%m-%d %H:%M:%S")

    def _setup_headers(self, ws: Worksheet) -> None:
        """Setup Excel headers with all case fields."""
        headers = [
            "ID",
            "Номер",
            "Номер производства",
            "Клиент",
            "ИНН клиента",
            "Email клиента",
            "Телефон клиента",
            "Орган",
            "Тип дела",
            "Тип объекта",
            "Адрес объекта",
            "Статус",
            "Дата начала",
            "Срок",
            "Дата завершения",
            "Стоимость",
            "Безналичный перевод",
            "Наличные",
            "Остаток долга",
            "Дебит (в пути)",
            "Истец",
            "Ответчик",
            "Судья",
            "Примечания",
            "Эксперты",
            "Создано",
            "Обновлено",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def _apply_formatting(self, ws: Worksheet) -> None:
        """Apply Excel formatting for better readability."""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
